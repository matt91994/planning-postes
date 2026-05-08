from flask import Flask, request, jsonify, send_file, make_response, session, redirect
from flask_cors import CORS
import openpyxl
import json
import io
import os
import psycopg2

app = Flask(__name__)
CORS(app)
app.secret_key = os.environ.get('SECRET_KEY', 'planning-secret-2026')

ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'planning2026')
DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql://postgres:g%403FPYtzL96%2C9dD@db.jdfnmpnmkhkvjalhvrhl.supabase.co:5432/postgres')

EMPLOYES = [{"absent":False,"id":4,"nom":"BRUNO","postesAutorises":[4,10,19,20]},{"absent":False,"id":5,"nom":"NENE","postesAutorises":[3,4,10,18]},{"absent":False,"id":6,"nom":"MAKAN","postesAutorises":[3,18,19,20]},{"absent":False,"id":7,"nom":"FREDDY","postesAutorises":[12,2]},{"absent":False,"id":8,"nom":"DALLA","postesAutorises":[2,12]},{"absent":False,"id":9,"nom":"JEAN SERVA","postesAutorises":[5,6,7,8,14,15,16,19,20,21,22,23,24,25,26]},{"absent":False,"id":10,"nom":"SYLLA","postesAutorises":[5,6,7,8,19,20]},{"absent":False,"id":11,"nom":"ROUSLAN","postesAutorises":[9,19,20]},{"absent":False,"id":12,"nom":"BOMOU","postesAutorises":[10,19,20,21,22,23,24,25,26]},{"absent":False,"id":13,"nom":"JEAN REMY","postesAutorises":[7,6,8,5,19,20]},{"absent":False,"id":14,"nom":"SIDY","postesAutorises":[5,6,7,8,16,19,20,21,22,23,24,25,26]},{"absent":False,"id":15,"nom":"ALEXIS","postesAutorises":[11]},{"absent":False,"id":16,"nom":"MADER","postesAutorises":[1]},{"absent":False,"id":17,"nom":"SYLVAIN LA SOMME","postesAutorises":[13,19,20,21,22,23,24,25,26]},{"absent":False,"id":18,"nom":"FISSIROU","postesAutorises":[14,15,19,20,21,22,23,24,25,26]},{"absent":False,"id":19,"nom":"ABDOULAYE","postesAutorises":[19,20,21,22,23,24,25,26]},{"absent":False,"id":20,"nom":"JEAN EDDY","postesAutorises":[14,15,16,19,20,21,22,23,24,25,26]},{"absent":False,"id":21,"nom":"DIAMBOU","postesAutorises":[15,14,21,22,23,24,25,26]},{"absent":False,"id":22,"nom":"CAMARA","postesAutorises":[5,6,7,8,11,14,15,16,19,20,21,22,23,24,25,26]},{"absent":False,"id":23,"nom":"CÉDRIC","postesAutorises":[16,17,18,19,20]},{"absent":False,"id":24,"nom":"SYLVAIN L","postesAutorises":[17,18]},{"absent":False,"id":25,"nom":"JEAN-MARC","postesAutorises":[21,22,23,24,25,26]},{"absent":False,"id":26,"nom":"TRAORÉ","postesAutorises":[21,22,23,24,25,26]},{"absent":False,"id":27,"nom":"RUI","postesAutorises":[13,21,14,15,22,23,24,25,26]},{"absent":False,"id":28,"nom":"SOULEYMANE","postesAutorises":[9,19,20,21,22,23,24,25,26]},{"absent":False,"id":29,"nom":"SAMOURA","postesAutorises":[19,20,21,22,23,24,25,26,2]},{"absent":False,"id":30,"nom":"DJIBRIL","postesAutorises":[19,20,21,22,23,24,25,26]},{"absent":False,"id":31,"nom":"VIRGIL","postesAutorises":[5,7,6,8,19,20,21,22,23,24,25,26,11]}]

POSTES = [{"cellule":"F23","id":1,"nom":"Déchargement","priorite":1},{"cellule":"F28","id":2,"nom":"Frigo","priorite":1},{"cellule":"I32","id":3,"nom":"Coupe droite","priorite":1},{"cellule":"I20","id":4,"nom":"Coupe gauche","priorite":1},{"cellule":"M32","id":5,"nom":"Lève droite 1","priorite":1},{"cellule":"Q32","id":6,"nom":"Lève droite 2","priorite":1},{"cellule":"M19","id":7,"nom":"Lève gauche 1","priorite":1},{"cellule":"P19","id":8,"nom":"Lève gauche 2","priorite":1},{"cellule":"U32","id":9,"nom":"Longe droite","priorite":1},{"cellule":"U19","id":10,"nom":"Longe gauche","priorite":1},{"cellule":"W28","id":11,"nom":"Bout du tapis","priorite":1},{"cellule":"S10","id":12,"nom":"Bout du tapis poitrine","priorite":1},{"cellule":"P15","id":13,"nom":"Poitrine","priorite":1},{"cellule":"M17","id":14,"nom":"BK doc1","priorite":1},{"cellule":"M15","id":15,"nom":"BK doc2","priorite":1},{"cellule":"M13","id":16,"nom":"BK commande","priorite":1},{"cellule":"K10","id":17,"nom":"Jambon","priorite":1},{"cellule":"H15","id":18,"nom":"Jambon aide","priorite":1},{"cellule":"U17","id":19,"nom":"Palette doc1","priorite":2},{"cellule":"U15","id":20,"nom":"Palette doc2","priorite":2},{"cellule":"X8","id":21,"nom":"Doc1","priorite":2},{"cellule":"X1","id":22,"nom":"Doc2","priorite":2},{"cellule":"AD8","id":23,"nom":"Doc3","priorite":2},{"cellule":"AD1","id":24,"nom":"Doc4","priorite":3},{"cellule":"AA1","id":25,"nom":"Doc5","priorite":3},{"cellule":"AA8","id":26,"nom":"Doc6","priorite":3}]

def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('CREATE TABLE IF NOT EXISTS config (key TEXT PRIMARY KEY, value TEXT NOT NULL)')
    cur.execute("SELECT value FROM config WHERE key='postes'")
    if not cur.fetchone():
        cur.execute("INSERT INTO config VALUES (%s,%s)", ('postes', json.dumps(POSTES)))
        cur.execute("INSERT INTO config VALUES (%s,%s)", ('employes', json.dumps(EMPLOYES)))
        cur.execute("INSERT INTO config VALUES (%s,%s)", ('nextEmpId', '32'))
        cur.execute("INSERT INTO config VALUES (%s,%s)", ('nextPostId', '27'))
    conn.commit()
    cur.close()
    conn.close()

init_db()

def is_logged_in():
    return session.get('logged_in') == True

LOGIN_PAGE = '''<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Planning — Connexion</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;600;700;800&display=swap" rel="stylesheet">
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: "DM Sans", sans-serif; background: #0d1117; min-height: 100vh; display: flex; align-items: center; justify-content: center; }
  .box { background: #161b22; border: 1px solid #30363d; border-radius: 16px; padding: 40px 36px; width: 100%; max-width: 380px; margin: 16px; }
  .logo-icon { font-size: 40px; text-align: center; margin-bottom: 10px; }
  h1 { color: #fff; font-size: 20px; font-weight: 800; text-align: center; }
  .subtitle { color: #7d8590; font-size: 13px; text-align: center; margin-top: 4px; margin-bottom: 28px; }
  label { display: block; font-size: 12px; font-weight: 700; color: #7d8590; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 6px; }
  input[type=password] { width: 100%; padding: 11px 14px; border-radius: 8px; background: #21262d; border: 1px solid #30363d; color: #e6edf3; font-size: 15px; font-family: "DM Sans", sans-serif; outline: none; margin-bottom: 16px; }
  input[type=password]:focus { border-color: #388bfd; }
  button { width: 100%; padding: 12px; border-radius: 8px; border: none; background: linear-gradient(90deg, #388bfd, #a371f7); color: #fff; font-size: 15px; font-weight: 700; cursor: pointer; font-family: "DM Sans", sans-serif; }
  .error { background: #3a111133; border: 1px solid #f8514966; border-radius: 8px; padding: 10px 14px; color: #fca5a5; font-size: 13px; margin-bottom: 16px; text-align: center; }
</style>
</head>
<body>
<div class="box">
  <div class="logo-icon">📋</div>
  <h1>Planning des Postes</h1>
  <p class="subtitle">Accès réservé à l équipe</p>
  {error}
  <form method="POST" action="/login">
    <label>Mot de passe</label>
    <input type="password" name="password" placeholder="••••••••" autofocus>
    <button type="submit">Se connecter</button>
  </form>
</div>
</body>
</html>'''

@app.route('/')
def index():
    if not is_logged_in():
        return make_response(LOGIN_PAGE.replace('{error}', ''), 200)
    try:
        with open(os.path.join(os.path.dirname(__file__), 'static', 'index.html'), 'r', encoding='utf-8') as f:
            html = f.read()
        resp = make_response(html)
        resp.headers['Content-Type'] = 'text/html; charset=utf-8'
        return resp
    except FileNotFoundError:
        return "Fichier index.html introuvable", 404

@app.route('/login', methods=['POST'])
def login():
    password = request.form.get('password', '')
    if password == ADMIN_PASSWORD:
        session['logged_in'] = True
        return redirect('/')
    error = '<div class="error">❌ Mot de passe incorrect</div>'
    return make_response(LOGIN_PAGE.replace('{error}', error), 401)

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')

@app.route('/config', methods=['GET'])
def get_config():
    if not is_logged_in():
        return jsonify({'error': 'Non autorisé'}), 401
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT key, value FROM config")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    data = {}
    for key, value in rows:
        try:
            data[key] = json.loads(value)
        except:
            data[key] = value
    return jsonify(data)

@app.route('/config', methods=['POST'])
def save_config():
    if not is_logged_in():
        return jsonify({'error': 'Non autorisé'}), 401
    try:
        data = request.get_json()
        conn = get_db()
        cur = conn.cursor()
        for key, value in data.items():
            v = json.dumps(value) if not isinstance(value, str) else value
            cur.execute("INSERT INTO config VALUES (%s,%s) ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value", [key, v])
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export', methods=['POST'])
def export():
    if not is_logged_in():
        return jsonify({'error': 'Non autorisé'}), 401
    try:
        planning = json.loads(request.form.get('planning'))
        excel_file = request.files.get('excel')
        wb = openpyxl.load_workbook(excel_file) if excel_file else openpyxl.Workbook()
        if not excel_file:
            wb.active.title = "Planning"
        ws = wb.active
        for item in planning:
            if item.get('cellule') and item.get('employe'):
                ws[item['cellule']] = item['employe']
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = excel_file.filename if excel_file else "planning.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
